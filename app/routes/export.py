from fastapi import APIRouter, Depends, BackgroundTasks
from fastapi.responses import FileResponse
import requests
import pandas as pd
import uuid
import os
from typing import List, Optional
from app.core.auth_utils import verificar_token

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from pydantic import BaseModel

router = APIRouter()

LOGO_PATH = "assets/Sublogo.png"

TEMP_DIR = "tmp"
os.makedirs(TEMP_DIR, exist_ok=True)

def remove_file(path: str):
    if os.path.exists(path):
        os.remove(path)

def get_nested_value(d: dict, path: str):
    keys = path.split(".")
    value = d
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key, None)
        else:
            value = None
        if value is None:
            break
    return value

class ExportColumn(BaseModel):
    field: str
    header: str
    width: int


class BodyExport(BaseModel):
    endpoint_url: Optional[str]
    excel_name: Optional[str] = 'planilha'
    method: Optional[str] = 'GET'
    columns: Optional[List[ExportColumn]]
    authorization: Optional[str]
    params: Optional[dict] = None
    body: Optional[dict] = None


@router.post("/api/v1/export-excel", tags=["Export"])
def export_excel(payload: BodyExport, background: BackgroundTasks, user_id: str = Depends(verificar_token)):
    endpoint_url = payload.endpoint_url
    excel_name = payload.excel_name or "export"
    columns = payload.columns
    token = payload.authorization
    params = payload.params or {}
    json_body = payload.body or {}

    headers = {}
    if token:
        headers["Authorization"] = token

    method = (payload.method or "GET").upper()
    if method == "GET":
        response = requests.get(endpoint_url, headers=headers, params=params, timeout=10)
    else:
        response = requests.post(endpoint_url, headers=headers, json=json_body, timeout=10)

    response.raise_for_status()
    data = response.json().get("data", [])

    if not data:
        return {"error": "Nenhum dado encontrado"}

    if columns:
        flattened_data = []
        for item in data:
            row = {
                col.header: get_nested_value(item, col.field)
                for col in columns
            }
            flattened_data.append(row)

        df = pd.DataFrame(flattened_data)
    else:
        df = pd.DataFrame(data)

    file_id = uuid.uuid4().hex
    file_path = f"{TEMP_DIR}/{excel_name}_{file_id}.xlsx"
    df.to_excel(file_path, index=False, startrow=7)

    wb = load_workbook(file_path)
    ws = wb.active

    if os.path.exists(LOGO_PATH):
        img = ExcelImage(LOGO_PATH)
        img.width = 180
        img.height = 120
        ws.add_image(img, "A1")

    if columns:
        for idx, col in enumerate(columns, start=1):
            ws.column_dimensions[
                ws.cell(row=8, column=idx).column_letter
            ].width = col.width or 15

    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)

    background.add_task(remove_file, file_path)

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{excel_name}.xlsx"
    )
